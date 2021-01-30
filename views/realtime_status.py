#!/usr/bin/python -tt
# -*- coding: utf-8 -*-

"""question.py
Define all api in realtime status pages
"""
from logging import getLogger
from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_from_directory
from routes import IsmsRouter
from schemas import *
from flask_sqlalchemy import SQLAlchemy
from libs.user_permission import admin_required
from libs.constants import CategoryTypeEnum, StatusTypeEnum
from config import LIMIT_CHOICE, LIMIT_ROWS
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os

db = SQLAlchemy()
logger = getLogger(__name__)
mod = Blueprint('realtime_status', __name__, url_prefix='/realtime_status', static_folder="static",
                template_folder="../templates/admin")


@mod.route('/exclude/<int:detail_id>')
@admin_required
def get_user_detail_exclude(detail_id):
    search = request.args.get('search')
    name = 'exclude'
    page = request.args.get('page', 1, type=int)
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_exclude = get_list_user_exclude_db(detail_id=detail_id, page=page, search=search)
    list_user_result_exclude = get_list_user_exclude(list_user_exclude=list_user_exclude.items)
    return render_template('realtime_status.html', SDKRouter=IsmsRouter,
                           list_user_result=list_user_result_exclude,
                           list_user_result_pagination=list_user_exclude,
                           detail=detail,
                           name_url=name)


@mod.route('/uncategorized/<int:detail_id>')
@admin_required
def get_user_detail_uncategorized(detail_id):
    search = request.args.get('search')
    name = 'uncategorized'
    page = request.args.get('page', 1, type=int)
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_took_the_exam_uncategorized = get_list_user_took_the_exam_uncategorized(detail_id=detail_id, page=page,
                                                                                      search=search)
    list_user_result_uncategorized = get_list_user_uncategorized(
        list_user_took_the_exam_uncategorized=list_user_took_the_exam_uncategorized.items)
    return render_template('realtime_status.html', SDKRouter=IsmsRouter,
                           list_user_result=list_user_result_uncategorized,
                           list_user_result_pagination=list_user_took_the_exam_uncategorized,
                           detail=detail,
                           name_url=name)


@mod.route('/fail/<int:detail_id>')
@admin_required
def get_user_detail_fail(detail_id):
    search = request.args.get('search')
    name = 'fail'
    page = request.args.get('page', 1, type=int)
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_took_the_exam = get_list_user_took_the_exam(detail_id=detail_id, status=False, page=page, search=search)
    list_user_result_fail = get_list_user_result_pass_or_fail(list_user_detail=list_user_took_the_exam.items,
                                                              status=False)
    return render_template('realtime_status.html', SDKRouter=IsmsRouter,
                           list_user_result=list_user_result_fail,
                           list_user_result_pagination=list_user_took_the_exam,
                           detail=detail,
                           name_url=name)


@mod.route('/pass/<int:detail_id>')
@admin_required
def get_user_detail_pass(detail_id):
    search = request.args.get('search')
    name = 'pass'
    page = request.args.get('page', 1, type=int)
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_took_the_exam = get_list_user_took_the_exam(detail_id=detail_id, status=True, page=page, search=search)
    list_user_result_pass = get_list_user_result_pass_or_fail(list_user_detail=list_user_took_the_exam.items,
                                                              status=True)
    return render_template('realtime_status.html', SDKRouter=IsmsRouter,
                           list_user_result=list_user_result_pass,
                           list_user_result_pagination=list_user_took_the_exam,
                           detail=detail,
                           name_url=name)


@mod.route('/<int:detail_id>/submit/<int:user_detail_id>', methods=['POST'])
@admin_required
def submit_force_status_user_detail(detail_id, user_detail_id):
    current_date = datetime.now()
    email = request.form['email']
    check_email_result = check_email(detail_id, email)
    if not check_email_result:
        email_model = Email(detail_id=detail_id, name=email)
        db.session.add(email_model)
        db.session.commit()
    status = request.form['status']
    if status == '1':
        status = StatusTypeEnum.FORCE_PASS.value
    else:
        status = StatusTypeEnum.FORCE_FAIL.value
    user_detail = db.session.query(Userdetail).filter(Userdetail.id == user_detail_id,
                                                      Userdetail.deleted_flag.isnot(True)).first()
    if not user_detail.start_time:
        db.session.query(Userdetail).filter(Userdetail.id == user_detail_id,
                                            Userdetail.deleted_flag.isnot(True)).update(
            dict(status_id=status, start_time=current_date, end_time=current_date))
    else:
        db.session.query(Userdetail).filter_by(id=user_detail_id).update(dict(status_id=status))
    db.session.commit()
    return 'success'


@mod.route('/export/exclude/<int:detail_id>')
@admin_required
def export_user_detail_exclude(detail_id):
    search = request.args.get('search')
    page = None
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_exclude = get_list_user_exclude_db(detail_id=detail_id, page=page, search=search)
    list_user_result_exclude = get_list_user_exclude(list_user_exclude=list_user_exclude)
    excel_file = os.path.join('static', 'templates_realtime_status.xlsx')
    wb = load_workbook(filename=excel_file)
    sheet_ranges = wb['Statistical']
    i = 1
    for item in list_user_result_exclude:
        i += 1
        sheet_ranges.cell(row=i, column=1).value = i - 1
        sheet_ranges.cell(row=i, column=2).value = item['email']
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=3).value = item['user_detail'].start_time
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=4).value = item['user_detail'].end_time - item['user_detail'].start_time
        sheet_ranges.cell(row=i, column=5).value = str(item['qms_total'] + item['isms_total']) + '/' + str(
            detail.qms_total + detail.isms_total)
        sheet_ranges.cell(row=i, column=6).value = str(item['qms_total']) + '/' + str(detail.qms_total)
        sheet_ranges.cell(row=i, column=7).value = str(item['isms_total']) + '/' + str(detail.isms_total)
    file_name = 'realtime_status_exam_exclude.xlsx'
    excel_file = os.path.join('static', file_name)
    wb.save(excel_file)
    return send_from_directory(directory='static', filename=file_name, as_attachment=True)


@mod.route('/export/uncategorized/<int:detail_id>')
@admin_required
def export_user_detail_uncategorized(detail_id):
    search = request.args.get('search')
    page = None
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_took_the_exam_uncategorized = get_list_user_took_the_exam_uncategorized(detail_id=detail_id, page=page,
                                                                                      search=search)
    list_user_result_uncategorized = get_list_user_uncategorized(
        list_user_took_the_exam_uncategorized=list_user_took_the_exam_uncategorized)
    excel_file = os.path.join('static', 'templates_realtime_status.xlsx')
    wb = load_workbook(filename=excel_file)
    sheet_ranges = wb['Statistical']
    i = 1
    for item in list_user_result_uncategorized:
        i += 1
        sheet_ranges.cell(row=i, column=1).value = i - 1
        sheet_ranges.cell(row=i, column=2).value = item['email']
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=3).value = item['user_detail'].start_time
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=4).value = item['user_detail'].end_time - item['user_detail'].start_time
        sheet_ranges.cell(row=i, column=5).value = str(item['qms_total'] + item['isms_total']) + '/' + str(
            detail.qms_total + detail.isms_total)
        sheet_ranges.cell(row=i, column=6).value = str(item['qms_total']) + '/' + str(detail.qms_total)
        sheet_ranges.cell(row=i, column=7).value = str(item['isms_total']) + '/' + str(detail.isms_total)
    file_name = 'realtime_status_exam_uncategorized.xlsx'
    excel_file = os.path.join('static', file_name)
    wb.save(excel_file)
    return send_from_directory(directory='static', filename=file_name, as_attachment=True)


@mod.route('/export/fail/<int:detail_id>')
@admin_required
def export_user_detail_fail(detail_id):
    search = request.args.get('search')
    page = None
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_took_the_exam = get_list_user_took_the_exam(detail_id=detail_id, status=False, page=page, search=search)
    list_user_result_fail = get_list_user_result_pass_or_fail(list_user_detail=list_user_took_the_exam,
                                                              status=False)
    excel_file = os.path.join('static', 'templates_realtime_status.xlsx')
    wb = load_workbook(filename=excel_file)
    sheet_ranges = wb['Statistical']
    i = 1
    for item in list_user_result_fail:
        i += 1
        sheet_ranges.cell(row=i, column=1).value = i - 1
        sheet_ranges.cell(row=i, column=2).value = item['email']
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=3).value = item['user_detail'].start_time
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=4).value = item['user_detail'].end_time - item['user_detail'].start_time
        sheet_ranges.cell(row=i, column=5).value = str(item['qms_total'] + item['isms_total']) + '/' + str(
            detail.qms_total + detail.isms_total)
        sheet_ranges.cell(row=i, column=6).value = str(item['qms_total']) + '/' + str(detail.qms_total)
        sheet_ranges.cell(row=i, column=7).value = str(item['isms_total']) + '/' + str(detail.isms_total)
    file_name = 'realtime_status_exam_fail.xlsx'
    excel_file = os.path.join('static', file_name)
    wb.save(excel_file)
    return send_from_directory(directory='static', filename=file_name, as_attachment=True)


@mod.route('/export/pass/<int:detail_id>')
@admin_required
def export_user_detail_pass(detail_id):
    search = request.args.get('search')
    page = None
    detail = db.session.query(Detail).filter(Detail.id == detail_id, Detail.deleted_flag.isnot(True)).first()
    list_user_took_the_exam = get_list_user_took_the_exam(detail_id=detail_id, status=True, page=page, search=search)
    list_user_result_pass = get_list_user_result_pass_or_fail(list_user_detail=list_user_took_the_exam,
                                                              status=True)
    excel_file = os.path.join('static', 'templates_realtime_status.xlsx')
    wb = load_workbook(filename=excel_file)
    sheet_ranges = wb['Statistical']
    i = 1
    for item in list_user_result_pass:
        i += 1
        sheet_ranges.cell(row=i, column=1).value = i - 1
        sheet_ranges.cell(row=i, column=2).value = item['email']
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=3).value = item['user_detail'].start_time
        if item['user_detail'].start_time:
            sheet_ranges.cell(row=i, column=4).value = item['user_detail'].end_time - item['user_detail'].start_time
        sheet_ranges.cell(row=i, column=5).value = str(item['qms_total'] + item['isms_total']) + '/' + str(
            detail.qms_total + detail.isms_total)
        sheet_ranges.cell(row=i, column=6).value = str(item['qms_total']) + '/' + str(detail.qms_total)
        sheet_ranges.cell(row=i, column=7).value = str(item['isms_total']) + '/' + str(detail.isms_total)
    file_name = 'realtime_status_exam_pass.xlsx'
    excel_file = os.path.join('static', file_name)
    wb.save(excel_file)
    return send_from_directory(directory='static', filename=file_name, as_attachment=True)


def check_email(detail_id, email):
    """
    Check if email exists in the email list to take the exam
    :param detail_id:
    :param email:
    :return: - True if email exists
             - False if email does not exist
    """
    list_email = db.session.query(Email.name).filter(Email.detail_id == detail_id, Email.deleted_flag.isnot(True)).all()
    email_name = list(sum(list_email, ()))
    return email in email_name


def get_list_user_took_the_exam(detail_id, status, page, search):
    """
    Get a list of users in the email list that took the exam
    :param detail_id:
    :return: list_user_detail
    """
    list_user_detail = db.session.query(Userdetail).join(User, Userdetail.user_id == User.id) \
        .join(Email, Email.detail_id == Userdetail.detail_id).join \
        (Maexamplestatu, Maexamplestatu.id == Userdetail.status_id) \
        .add_columns(User.email, Maexamplestatu.status).filter(
        Userdetail.detail_id == detail_id,
        Userdetail.end_time,
        Email.name == User.email).filter(Maexamplestatu.status == status).filter(Userdetail.deleted_flag.isnot(True),
                                                                                 User.deleted_flag.isnot(True),
                                                                                 Email.deleted_flag.isnot(True),
                                                                                 Maexamplestatu.deleted_flag.isnot(
                                                                                     True))
    if search:
        list_user_detail = list_user_detail.filter(Email.name.like('%' + search + '%'))
    if page:
        return list_user_detail.order_by(Userdetail.end_time.desc()).paginate(page, per_page=LIMIT_ROWS)
    return list_user_detail.order_by(Userdetail.end_time.desc()).all()


def get_list_user_took_the_exam_uncategorized(detail_id, page, search):
    """
    Get a list of users in the email list that took the exam
    :param detail_id
    :return: list_user_detail
    """
    list_email_db = db.session.query(Email.name).filter_by(detail_id=detail_id).all()
    list_email = list(sum(list_email_db, ()))
    list_user_detail_uncategorized = db.session.query(Userdetail).join(User, Userdetail.user_id == User.id) \
        .join(Email, Email.detail_id == Userdetail.detail_id).join \
        (Maexamplestatu, Maexamplestatu.id == Userdetail.status_id) \
        .add_columns(User.email, Maexamplestatu.status).filter(
        Userdetail.detail_id == detail_id,
        Userdetail.end_time).filter(User.email.notin_(list_email)).filter(Userdetail.deleted_flag.isnot(True),
                                                                          User.deleted_flag.isnot(True),
                                                                          Email.deleted_flag.isnot(True),
                                                                          Maexamplestatu.deleted_flag.isnot(
                                                                              True)).group_by(Userdetail.id)
    if search:
        list_user_detail_uncategorized = list_user_detail_uncategorized.filter(Email.name.like('%' + search + '%'))
    if page:
        return list_user_detail_uncategorized.order_by(Userdetail.end_time.desc()).paginate(page, per_page=LIMIT_ROWS)
    return list_user_detail_uncategorized.order_by(Userdetail.end_time.desc()).all()


def get_list_user_not_took_the_exam(detail_id, list_user_took_the_exam):
    """
    Get a list email of users in the email list who did not take the exam
    :param detail_id:
    :param list_user_took_the_exam:
    :return: list_email
    """
    list_email_db = db.session.query(Email.name).filter(Email.detail_id == detail_id,
                                                        Email.deleted_flag.isnot(True)).all()
    list_email = list(sum(list_email_db, ()))
    for item in list_user_took_the_exam:
        list_email.remove(item.email)
    return list_email


def get_list_question_pass_of_user_detail_by_category(user_detail_id, category):
    """
    Get a list questions pass of users by category
    :param user_detail_id: 
    :param category: 
    :return: list_user_question
    """
    return db.session.query(Userquestion).join(Question, Question.id == Userquestion.question_id). \
        join(Answer, Answer.id == Userquestion.answer_id).add_columns(Question.category).filter(
        Userquestion.user_detail_id == user_detail_id, Answer.is_correct, Question.category == category).all()


def get_list_user_result_pass_or_fail(list_user_detail, status):
    """
    Converts data list user pass or fail into statistical data result format
    :param list_user_detail:
    :return: list_user_result
    """
    list_user_result = []
    for item in list_user_detail:
        list_question_qms = get_list_question_pass_of_user_detail_by_category(item.Userdetail.id,
                                                                              CategoryTypeEnum.QMS_Category.value)
        list_question_isms = get_list_question_pass_of_user_detail_by_category(item.Userdetail.id,
                                                                               CategoryTypeEnum.ISMS_Category.value)
        if item.status == status:
            list_user_result.append(dict(
                user_detail=item.Userdetail,
                email=item.email,
                qms_total=len(list_question_qms),
                isms_total=len(list_question_isms),
                status=item.status
            ))
    return list_user_result


def get_list_user_uncategorized(list_user_took_the_exam_uncategorized):
    """
    Converts data list user uncategorized into statistical data result format
    :param list_user_took_the_exam_uncategorized:
    :return: list user result
    """
    list_user_result = []
    for item in list_user_took_the_exam_uncategorized:
        list_question_qms = get_list_question_pass_of_user_detail_by_category(item.Userdetail.id,
                                                                              CategoryTypeEnum.QMS_Category.value)
        list_question_isms = get_list_question_pass_of_user_detail_by_category(item.Userdetail.id,
                                                                               CategoryTypeEnum.ISMS_Category.value)
        list_user_result.append(dict(
            user_detail=item.Userdetail,
            email=item.email,
            qms_total=len(list_question_qms),
            isms_total=len(list_question_isms),
            status=item.status
        ))
    return list_user_result


def get_list_user_exclude(list_user_exclude):
    """
    Converts data list user exclude into statistical data result format
    :param list_user_exclude:
    :return: lít_user_result
    """
    list_user_result = []
    for item in list_user_exclude:
        list_user_result.append(dict(
            user_detail=item.Userdetail,
            email=item.email,
            qms_total=0,
            isms_total=0,
            status=item.status
        ))
    return list_user_result


def get_list_user_exclude_db(detail_id, page, search):
    """
    Get list user exclude in database
    :param detail_id:
    :param page:
    :return: list_user_detail
    """
    list_user_exclude = db.session.query(Userdetail).join(User, User.id == Userdetail.user_id).join \
        (Maexamplestatu, Maexamplestatu.id == Userdetail.status_id).join(Email,
                                                                         Email.name == User.email).filter(
        Userdetail.detail_id == detail_id, Userdetail.start_time == None, Userdetail.deleted_flag.isnot(True),
        User.deleted_flag.isnot(True), Maexamplestatu.deleted_flag.isnot(True), Email.deleted_flag.isnot(True)) \
        .add_columns(User.email, Maexamplestatu.status).group_by(Userdetail.id)
    if search:
        list_user_exclude = list_user_exclude.filter(User.email.like('%' + search + '%'))
    if page:
        return list_user_exclude.order_by(
            Userdetail.updated_at.desc()).paginate(page, per_page=LIMIT_ROWS)
    return list_user_exclude.order_by(
        Userdetail.updated_at.desc()).all()
