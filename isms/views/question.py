#!/usr/bin/python -tt
# -*- coding: utf-8 -*-

"""question.py
Define all api in question pages
"""
import os
import json
import unidecode
from logging import getLogger

import pandas as pd
from flask_sqlalchemy import SQLAlchemy
from flask import Blueprint, render_template, request, redirect, url_for, send_from_directory
from routes import IsmsRouter
from libs.user_permission import admin_required
from config import LIMIT_CHOICE, LIMIT_ROWS
from schemas import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

db = SQLAlchemy()
logger = getLogger(__name__)
mod = Blueprint('question', __name__, url_prefix='/question', static_folder="static",
                template_folder="../templates/admin")


@mod.route('/')
@admin_required
def get_all_question():
    """
        Get all question
        Search question by content
    """
    search = request.args.get('search')
    page = request.args.get('page', 1, type=int)
    question_list = db.session.query(Question).filter(Question.deleted_flag.isnot(True))
    if search:
        question_list = question_list.filter(Question.content.like('%' + search + '%'))
    if page < 1:
        return redirect(url_for('question.get_all_question'))
    question_list = question_list.order_by(Question.updated_at.desc()).paginate(page, per_page=LIMIT_ROWS)
    return render_template('question_list.html', SDKRouter=IsmsRouter, questions=question_list)


@mod.route('/create', methods=['GET', 'POST'])
@admin_required
def create_question():
    """
        Create question
    """
    if request.method == 'POST':
        content_question = request.form['content_question']
        category = request.form['category']
        answers = request.form['answers']
        answers = json.loads(answers)
        existed_question = check_question_exists(question_str=content_question)
        logger.info(f"Question {content_question} is existed!")
        question = Question(content=content_question, category=category)
        db.session.add(question)
        db.session.commit()
        answers_sql = []
        for item in answers:
            answers_sql.append(Answer(question_id=question.id,
                                      content=item['content'],
                                      is_correct=item['is_correct']))
        db.session.bulk_save_objects(answers_sql)
        db.session.commit()
        return redirect(url_for('question.get_all_question'))
    return render_template('create_question.html', SDKRouter=IsmsRouter, limit_choice=LIMIT_CHOICE)


@mod.route('/update/<int:question_id>', methods=['GET', 'POST'])
@admin_required
def update_question(question_id):
    """
        Get question by ID
        Update question by ID
        And update answer by question ID
    """
    if request.method == 'POST':
        content_question = request.form['content_question']
        category = request.form['category']
        answers = request.form['answers']
        answers = json.loads(answers)
        new_question = Question(id=question_id, content=content_question, category=category)
        db.session.merge(new_question)
        db.session.query(Answer).filter(Answer.question_id == question_id).delete()
        db.session.commit()
        answers_sql = []
        for item in answers:
            answers_sql.append(Answer(question_id=question_id,
                                      content=item['content'],
                                      is_correct=item['is_correct']))
        db.session.bulk_save_objects(answers_sql)
        db.session.commit()
        # TODO use merge answer
        return redirect(url_for('question.get_all_question'))
    question = db.session.query(Question).filter(Question.id == question_id).first()
    answer_list = db.session.query(Answer).filter(Answer.question_id == question_id).all()
    return render_template('update_question.html', SDKRouter=IsmsRouter, limit_choice=LIMIT_CHOICE, question=question,
                           answer_list=answer_list)


@mod.route('/import', methods=['POST'])
@admin_required
def import_question():
    if request.method == 'POST':
        f = request.files['file']
        data_xls = pd.read_excel(f)
        for item, row in data_xls.iterrows():
            question = row['Câu hỏi']
            category = row['Thể loại']
            answer = row['Phương án']
            result_check_question = check_question_exists(question.strip())
            if not result_check_question:
                question = Question(content=question, category=category)
                db.session.add(question)
                db.session.commit()
                list_answer = process_the_answer(answer_str=answer, question_id=question.id)
                db.session.bulk_save_objects(list_answer)
                db.session.commit()
        return redirect(url_for('question.get_all_question'))


@mod.route('/export')
@admin_required
def export_question():
    excel_file = os.path.join('static', 'templates_question.xlsx')
    wb = load_workbook(filename=excel_file)
    sheet_ranges = wb['Question']
    list_question_db = db.session.query(Question).filter(Question.deleted_flag.isnot(True)).all()
    list_answer_str = []
    list_result_str = []
    for item in list_question_db:
        list_answer_db = item.answers
        list_answer_str.append(get_answer_str(list_answer_db))
        list_result_str.append(get_list_result_str(list_answer_db))
        x = 1
    list_answer_db = db.session.query(Answer.content).filter(Answer.deleted_flag.isnot(True)).all()
    # list_question = list(sum(list_question_db, ()))
    i = 1
    for item in list_question_db:
        i += 1
        sheet_ranges.cell(row=i, column=1).value = i - 1
        sheet_ranges.cell(row=i, column=2).value = item.content
        sheet_ranges.cell(row=i, column=4).value = item.category
    i = 1
    for item in list_answer_str:
        i += 1
        sheet_ranges.cell(row=i, column=3).value = item
        sheet_ranges.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
    i = 1
    for item in list_result_str:
        i += 1
        sheet_ranges.cell(row=i, column=5).value = item
    file_name = 'question.xlsx'
    excel_file = os.path.join('static', file_name)
    wb.save(excel_file)
    return send_from_directory(directory='static', filename=file_name, as_attachment=True)


@mod.route('/<int:question_id>')
@admin_required
def delete_question_by_id(question_id):
    """
        Delete question by ID
        Soft remove question, enable deleted_flag
    """
    db.session.query(Question).filter_by(id=question_id).update(dict(deleted_flag=1))
    db.session.commit()
    return 'Delete question successfully'


def process_the_answer(answer_str, question_id):
    """
    Process the sequence of answers
    :param answer_str:
    :return: list answer
    """
    list_answer = []
    list_answer_and_index = answer_str.split('\n')
    first_choice = True
    for index, item in enumerate(list_answer_and_index):
        list_answer.append(
            Answer(content=item.split('.', 1)[1].strip(), question_id=question_id, is_correct=first_choice))
        first_choice = False
    return list_answer


def check_question_exists(question_str):
    """
    Check question exists out system
    :param question_str:
    :return: True or False
    """
    question = db.session.query(Question.content_plain).filter(Question.content_plain == question_str).first()
    if question:
        return True
    return False


@mod.route('/test', methods=['GET', 'POST'])
@admin_required
def test_create_question():
    """
        Create question
    """
    content_question = """
       Trước cửa ra/vào của phòng làm vi ệc có dán biển: Khu vực hạn chế - Không sự dụng máy ảnh. Trưa hôm đó, một nhân viên trong phòng ?
    """

    # Change vietnamese to latin
    content_question = unidecode.unidecode(content_question).strip()

    # Avoid SQL injection
    content_question = content_question.replace('"', "")
    logger.info(f"content: {content_question}")
    existed_question = check_question_exists(question_str=content_question)
    raw_query_sql = f"""
        SELECT id, content FROM question
            WHERE MATCH (content_plain) AGAINST ("{content_question}" IN NATURAL LANGUAGE MODE)
            ORDER BY MATCH (content_plain) AGAINST ("{content_question}" IN NATURAL LANGUAGE MODE) DESC LIMIT 1;
    """
    result = db.session.execute(raw_query_sql)
    for row in result:
        logger.info(f"row = {row.id}, content = {row.content}")
    if existed_question:
        logger.info(f"Question {content_question} is existed!")
        return "Existed"
    else:
        return "Not-Existed"


def get_answer_str(list_answer):
    """
    Get chain of answers
    :param list_answer:
    :return:
    """
    list_content_answer = []
    for index, item in enumerate(list_answer):
        list_content_answer.append(index_upper_alpha(index) + ". " + item.content)
    str_one = '\r\n'
    str_two = list_content_answer
    return str_one.join(str_two)


def get_list_result_str(list_answer):
    """
    Get result answer
    :param list_answer:
    :return: str
    """
    result_answer = ''
    for index, item in enumerate(list_answer):
        if item.is_correct:
            result_answer = index_upper_alpha(index)
    return result_answer


def index_upper_alpha(i):
    """
    Get index alpha
    :param i:
    :return: str
    """
    switcher = {
        0: 'A',
        1: 'B',
        2: 'C',
        3: 'D',
        4: 'E',
        5: 'F'
    }
    return switcher.get(i, "Invalid day of week")
